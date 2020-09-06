#!/usr/bin/python
# -*- coding: utf-8 -*-

############################
# **** IMPORT SECTION **** #
############################
# стандартные пакеты
import sys, os, configparser

# пакеты linuxcnc
import linuxcnc, hal # http://linuxcnc.org/docs/html/hal/halmodule.html

# пакеты GUI
from PyQt5 import QtCore, QtWidgets
#from qtvcp.widgets import FocusOverlay
from qtvcp.widgets.mdi_line import MDILine as MDI_WIDGET
from qtvcp.widgets.gcode_editor import GcodeEditor as GCODE
from qtvcp.lib.keybindings import Keylookup
from qtvcp.core import Status, Action

# пакеты для графиков
import pyqtgraph as pg
from pyqtgraph import PlotWidget, plot

# пакет для xlsx
import openpyxl

# Set up logging
from qtvcp import logger

LOG = logger.getLogger(__name__)

# Set the log level for this module
#LOG.setLevel(logger.INFO) # One of DEBUG, INFO, WARNING, ERROR, CRITICAL

###########################################
# **** INSTANTIATE LIBRARIES SECTION **** #
###########################################

KEYBIND = Keylookup()
STATUS = Status()
ACTION = Action()
###################################
# **** HANDLER CLASS SECTION **** #
###################################

class HandlerClass:

    ########################
    # **** INITIALIZE **** #
    ########################
    # widgets allows access to  widgets from the qtvcp files
    # at this point the widgets and hal pins are not instantiated
    def __init__(self, halcomp,widgets,paths):
        self.hal = halcomp
        self.w = widgets
        self.PATHS = paths
        
        #self.siggen_test_read_pin = self.hal.newpin('siggen_test_read_pin',
        #                                            hal.HAL_FLOAT, hal.HAL_IN) # тестовый пин для получения синусоиды с siggen http://linuxcnc.org/docs/2.8/html/hal/halmodule.html#_use_with_hal_glib_in_qtvcp_handler

        #self.siggen_test_read_pin.value_changed.connect(lambda s: self.on_siggen_test_read_pin_value_changed(s)) # connect the pin to a callback http://linuxcnc.org/docs/2.8/html/hal/halmodule.html#_use_with_hal_glib_in_qtvcp_handler

    ##########################################
    # SPECIAL FUNCTIONS SECTION              #
    ##########################################

    # at this point:
    # the widgets are instantiated.
    # the HAL pins are built but HAL is not set ready
    # This is where you make HAL pins or initialize state of widgets etc
    def initialized__(self):
        database_file = 'База данных испытаний.xlsx'
        try:
            wb = openpyxl.load_workbook(filename = database_file, read_only = True)
            print("*** OK load_workbook")
            for sheet in wb:
                print(sheet.title)
            ws = wb[wb.sheetnames[1]]
            print("*** OK wb[\"Modeli\"]")
            list_test_types = []
            row = 1
            print("*** OK value 1", ws.cell(row = 1, column = 1).value)
            print("*** OK value 2", ws.cell(row = 2, column = 1).value)
            print("*** OK value 3", ws.cell(row = 3, column = 1).value)
            
            while(ws.cell(row = row, column = 1).value is not None):
                list_test_types.append(ws.cell(row = row, column = 1).value)
                print("*** OK row=", row, " appended, value = ",
                      ws.cell(row = row, column = 1).value)
                row = row + 1
            print("*** list_test_types = ", list_test_types) 
            self.w.cmbModel.addItems(list_test_types)
            print("*** OK ", addItems)
        except:
            print("*** Не удалось открыть файл базы данных \"", database_file, "\"")
            #fov = FocusOverlay(self)
            #fov.show()
        self.params_and_controls_dict = (
            {'GEAR': self.w.spnGear21,
            'PITCH' : self.w.spnPitch21,
            'TRAVEL' : self.w.spnTravel21,
            'NOM_VEL' : self.w.spnNom_Vel21,
            'NOM_ACCEL' : self.w.spnNom_Accel21},
            {'GEAR' : self.w.spnGear22,
            'NOM_VEL' : self.w.spnNom_Vel22,
            'BRAKE_TORQUE' : self.w.spnBrake_Torque22,
            'DURATION' : self.w.spnDuration22},
            {'GEAR' : self.w.spnGear23,
            'PITCH' : self.w.spnPitch23,
            'NOM_DSP_IDLE' : self.w.spnNom_Dsp_Idle23,
            'NOM_VEL_IDLE' : self.w.spnNom_Vel_Idle23,
            'NOM_ACCEL_IDLE' : self.w.spnNom_Accel_Idle23,
            'NOM_DSP_MEASURE' : self.w.spnNom_Dsp_Measure23,
            'NOM_VEL_MEASURE' : self.w.spnNom_Vel_Measure23,
            'NOM_ACCEL_MEASURE' : self.w.spnNom_Accel_Measure23,
            'NOM_LOAD' : self.w.spnNom_Load23,
            'NOM_POS_MEASURE' : self.w.spnNom_Pos_Measure23},
            {'GEAR' : self.w.spnGear24,
            'PITCH' : self.w.spnPitch24,
            'NOM_TRAVEL' : self.w.spnNom_Travel24,
            'NOM_OMEGA' : self.w.spnNom_Omega24,
            'NOM_ACCEL_COEFF' : self.w.spnNom_Accel_Coeff24,
            'NOM_F1' : self.w.spnNom_F1_24,
            'NOM_F2' : self.w.spnNom_F2_24,
            'OVERLOAD_COEFF' : self.w.spnOverload_Coeff24,
            'DWELL' : self.w.spnDwell24,
            'N' : self.w.spnN24,
            'LOG_FREQ' : self.w.spnLog_Freq24} )

    ########################
    # CALLBACKS FROM STATUS#
    ########################

    #######################
    # CALLBACKS FROM FORM #
    #######################
    
    def onBtnNext1(self):
        print "*** onBtnNext clicked"
        self.DRIVE = self.w.cmbDrive1.currentIndex()
        self.NAME = self.w.edtName1.text()
        self.DATE = self.w.edtDate1.text()
        self.PART = self.w.edtPart1.text()
        self.MODEL = self.w.cmbModel1.currentText()
        if(self.w.cmbType1.currentIndex()<4):
            self.load_ini(self.w.cmbType1.currentIndex())
            self.w.stackedWidget.setCurrentIndex(self.w.cmbType1.currentIndex()+1)
        else:
            self.load_ini(3)
            self.w.stackedWidget.setCurrentIndex(3)
        # self.close() только для случая многооконного интерфейса
    
    def onBtnTempShow1(self):
        self.w.stackedWidget.setCurrentIndex(0)
        pass
        
    def onBtnTempShow21(self):
        self.w.stackedWidget.setCurrentIndex(1)
        pass
        
    def onBtnTempShow22(self):
        self.w.stackedWidget.setCurrentIndex(2)
        pass
        
    def onBtnTempShow23(self):
        self.w.stackedWidget.setCurrentIndex(3)
        pass
        
    def onBtnTempShow24(self):
        self.w.stackedWidget.setCurrentIndex(4)
        pass
        
    def onBtnTempShow31(self):
        self.w.stackedWidget.setCurrentIndex(5)
        pass
        
    def onBtnTempShow32(self):
        self.w.stackedWidget.setCurrentIndex(6)
        pass
        
    def onBtnTempShow33(self):
        self.w.stackedWidget.setCurrentIndex(7)
        pass
        
    def onBtnTempShow34(self):
        self.w.stackedWidget.setCurrentIndex(8)
        pass

    #####################
    # GENERAL FUNCTIONS #
    #####################
    
    def on_siggen_test_read_pin_value_changed(self, data):
        return
        #print("*** siggen pin data: ", data)
        #print("*** siggen pin: ", self.siggen_test_read_pin.get())
        #print("*** siggen.0.sine directly", hal.get_value("siggen.0.sine"))

    def load_ini(self, n_form):
         # открытие и парсинг INI-файла, соответствующего типу испытания, в объект config
        config = configparser.ConfigParser()
        config.read('BallScrewControl' + str(n_form+1) + '.ini')
        # занесение значений из config в интерфейс по словарю
        for key in self.params_and_controls_dict[n_form]:
            self.params_and_controls_dict[n_form][key].setValue(float(config['PARAMETERS'][key]))
        #TODO обработка ошибок и исключений: 1) нет файла - сообщение и заполнение по умолчанию, создание конфига
        #TODO обработка ошибок и исключений: 2) нет ключей в конфиге - сообщение и заполнение по умолчанию

    def save_ini(self, n_form):
        # перемещение значений  из интерфейса в объект config
        config = configparser.ConfigParser()
        for key in self.params_and_controls_dict[n_form]:
            config['PARAMETERS'][key] = str(self.params_and_controls_dict[n_form][key].value())

        # запись заполненного config в ini-файл
        with open('BallScrewControl' + str(n_form+1) + '.ini', 'w') as configfile:
            config.write(configfile)

        #TODO обработка ошибок и исключений
        return

    
################################
# required handler boiler code #
################################

def get_handlers(halcomp, widgets, paths):
     return [HandlerClass(halcomp, widgets, paths)]
