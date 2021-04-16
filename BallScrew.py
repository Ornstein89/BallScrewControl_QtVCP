# This Python file uses the following encoding: utf-8
#!/usr/bin/python
# -*- coding: utf-8 -*-

import sys, os, codecs, io
import configparser
from shutil import copyfile

from PyQt5 import uic
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QPushButton
from PyQt5.QtWidgets import QLayout, QSizePolicy
from PyQt5.QtCore import QFile, QDateTime, QTime
from PyQt5.QtGui import QIcon

# пакет для xlsx
import openpyxl

class M(dict):
    def __setitem__(self, key, value):
        if key in self:
            items = self[key]
            print 'value=', value
            new = value[0]
            if new not in items:
                items.append(new)
        else:
            super(M, self).__setitem__(key, value)

class BallScrew(QMainWindow):
    def __init__(self):
        super(BallScrew, self).__init__()
        uic.loadUi("BallScrew.ui", self)
        self.setWindowIcon(QIcon("BallScrewControlIcon.png"))
        self.initParamsGUIMatch()
        self.loadIni()
        self.initGUI()

        #self.load_ui()

    def onBtnFileSelect(self):
        fname = QFileDialog.getSaveFileName(self, 'Save log file',
        ".","LOG (*.log);;Text (*.txt);;All Files (*.*)")
        #TODO не выбран
        self.LOGFILE = fname[0]
        self.edtFileSelect.setText(self.LOGFILE)

    def onBtnNext1(self):
        print "*** onBtnNext clicked"
        self.wgtNext2.show()
        self.DRIVE = self.cmbDrive.currentIndex()
        self.NAME = self.edtName.text().encode('utf-8')
        self.DATE = self.edtDate.text()
        self.PART = self.edtPart.text()
        self.MODEL = self.cmbModel.currentText()
        if (self.cmbType.currentIndex()<0):
            pass
        elif(self.cmbType.currentIndex()<4):
            self.TYPE_N = self.cmbType.currentIndex()+1
            print "*** currentIndex() = ", self.cmbType.currentIndex()
            print "*** self.TYPE_N = ", self.TYPE_N
        else:
            self.TYPE_N = 4
        self.set_default_values(self.TYPE_N)
        self.edtFileSelect.setText(self.MODEL + '__' + self.PART + '__'
        + self.TYPE + '__' + self.DATE + '__.log')
        self.stackedWidget.setCurrentIndex(self.TYPE_N)
        self.checkGUI2()
        self.onEdtFileChanged()

    def onBtnNext2(self):
        #запись ini-файла
        self.onBtnSaveIni()
        #запуск linuxcnc с ini-файлом
        #os.system(u"linuxcnc ./"+self.NAME.decode('utf-8'))
        #os.system(u"linuxcnc ./"+self.NAME.encode('utf-8')) # вариант опробовать
        # вариант опробовать os.system(u"linuxcnc ./"+unicode(self.NAME))
        os.system(("linuxcnc ./" + self.CNCINIFILENAME).encode('utf-8')) # вариант опробовать
        return

    def onBtnBack(self):
        self.wgtNext2.hide()
        self.stackedWidget.setCurrentIndex(0)
        self.checkGUI1()

    def onForm1DataChanged(self):
        self.MODEL = self.cmbModel.currentText()
        self.PART = self.edtPart.text()
        self.TYPE = self.cmbType.currentText().replace(' ', '_')
        self.DATE = self.edtDate.text()
        self.edtName.setText(self.MODEL + '__' + self.PART + '__'
                        + self.TYPE + '__' + self.DATE)#### + '__.ini')
        self.checkGUI1()

    def onEdtFileChanged(self):
        self.LOGFILE = self.edtFileSelect.text().encode('utf-8')
        print '***self.LOGFILE=', self.LOGFILE
        self.checkGUI2()

    def onBtnDefaultAll(self):
        self.set_default_values(self.TYPE_N)
        pass

    def initGUI(self):
        #видимость/невидимость, активность/неактивность
        self.stackedWidget.setCurrentIndex(0)
        self.btnNext1.setEnabled(False)
        self.btnNext2.setEnabled(False)
        self.wgtNext2.setVisible(False)
        self.wgtTest.setVisible(False)
        self.add_default_buttons()
        self.loadExcelFile()
        self.edtDate.setDateTime( QDateTime.currentDateTime())
        self.checkGUI1()

    def checkGUI1(self):  # проверка состояния формы 1 и разрешение перехода на форму 2
        canNext = ( (self.MODEL != "")
                    and (self.cmbDrive.currentIndex >= 0)
                    and (self.TYPE != "")
                    and (self.PART != ""))
        #TODO перенастроить на переменные self.TYPE и др.
        if canNext:
            self.btnNext1.setEnabled(True)
        else:
            self.btnNext1.setEnabled(False)
            #TODO уведомление "Заполните поля корректными значениями"

    def checkGUI2(self): # проверка состояния формы 2 и разрешение запуска linuxcnc (форма 3)
        canNext2 = self.edtFileSelect.text().endswith(".log")
        if canNext2:
            self.btnNext2.setEnabled(True)
        else:
            self.btnNext2.setEnabled(False)

    def loadExcelFile(self): #TODO переписать когда будет выдана библиотека в excel
        database_file = 'База данных испытаний.xlsx'
        try:
            wb = openpyxl.load_workbook(filename = database_file, read_only = True)
            print("*** OK load_workbook")
            #for sheet in wb:
            #    print(sheet.title)
            ws = wb[wb.sheetnames[1]]
            print("*** OK wb[\"Modeli\"]")
            list_test_types = []
            row = 1
            print("*** OK value 1", ws.cell(row = 1, column = 1).value)
            print("*** OK value 2", ws.cell(row = 2, column = 1).value)
            print("*** OK value 3", ws.cell(row = 3, column = 1).value)

            while(ws.cell(row = row, column = 1).value is not None):
                list_test_types.append(ws.cell(row = row, column = 1).value)
                print("*** OK row=", row, " appended, value = ",
                      ws.cell(row = row, column = 1).value)
                row = row + 1
            print("*** list_test_types = ", list_test_types)
            self.cmbModel.addItems(list_test_types)
            self.cmbModel.setCurrentIndex(-1)
            print("*** OK ", addItems)
        except:
            print("*** Не удалось открыть файл базы данных \"", database_file, "\"")

    def loadIni(self):
        self.config = configparser.ConfigParser(strict=False) # strict=False - разрешение повторов имени ключа
        self.config.optionxform = str # имена ключей не будут приведены к нижнему регистру
        self.config.read('BallScrew.ini')

    def add_default_buttons(self): #создать рядом с элементом управления кнопку сброса
        #assert 0 < n_form < 5, "Номер формы (типа испытания) не входит в пределы 1..4 и равен " + str(n_form)
        for n_form in range(1, 5):
            for key, control in self.params_and_controls_dict[n_form-1].items():
                val_string = self.config['TYPE_'+str(n_form)][key]
                parent = control.parentWidget()
                layout = control.parentWidget().layout()
                index = layout.indexOf(control)
                #print "key = ", key, "; index = ", index
                button = QPushButton(val_string, parent)
                #button.setMaximumSize(control.size())
                button.setMaximumHeight(control.height())
                button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
                location = control.parentWidget().layout().getItemPosition(index) #TODO здесь возникает ошибка!!
                control.parentWidget().layout().addWidget(button, location[0], location[1]+1)
                if (control.__class__.__name__ == 'QTimeEdit'):
                    button.clicked.connect(
                        lambda state, p_control=control, v=QTime(0,0,0).addSecs(int(val_string)): p_control.setTime(v))
                else:
                    button.clicked.connect(
                        lambda state, p_control=control, v=float(val_string): p_control.setValue(v))

    def set_default_values(self, n_form):
        assert 0 < n_form < 5, "Номер формы (типа испытания) не входит в пределы 1..4 и равен " + str(n_form)
        for key, control in self.params_and_controls_dict[n_form-1].items():
            if 'TYPE_'+str(n_form) not in self.config:
                return
            val_string = self.config['TYPE_'+str(n_form)][key]
            print "*** set_default_values(): control.__class__.__name__ = ", control.__class__.__name__
            if (control.__class__.__name__ == 'QTimeEdit'): # если QTimeEdit - перевести значения из секунд
                print "*** val_string=", val_string

                maxTime = QTime(0,0,0).addSecs(int(self.config['TYPE_'+str(n_form)][key+'_MAX']));
                print "*** max time=", maxTime
                control.setMaximumTime(maxTime)

                minTime = QTime(0,0,0).addSecs(int(self.config['TYPE_'+str(n_form)][key+'_MIN']));
                control.setMinimumTime(minTime)

                tmpTime = QTime(0,0,0).addSecs(int(val_string))
                control.setTime(tmpTime)
                #control.setMinimum(float(self.config['TYPE_'+str(n_form)][key+'_MIN']))
                #control.setMaximum(float(self.config['TYPE_'+str(n_form)][key+'_MAX']))
                #control.setValue(float(val_string))
            else: # если обычный QEdit
                control.setMinimum(float(self.config['TYPE_'+str(n_form)][key+'_MIN']))
                control.setMaximum(float(self.config['TYPE_'+str(n_form)][key+'_MAX']))
                control.setValue(float(val_string))

    def initParamsGUIMatch(self):
        self.params_and_controls_dict = (
            {'GEAR': self.spnGear21,
            'PITCH' : self.spnPitch21,
            'TRAVEL' : self.spnTravel21,
            'NOM_VEL' : self.spnNom_Vel21,
            'NOM_ACCEL' : self.spnNom_Accel21,
            'MICROSTEP' : self.spnMicrostep21},

            {'GEAR' : self.spnGear22,
            'NOM_VEL' : self.spnNom_Vel22,
            'BRAKE_TORQUE' : self.spnBrake_Torque22,
            'DURATION' : self.timeEdit, # self.spnDuration22,
            'EFFICIENCY' : self.spnEfficiency22,
            'MICROSTEP' : self.spnMicrostep22},

            {'GEAR' : self.spnGear23,
            'PITCH' : self.spnPitch23,
            'NOM_DSP_IDLE' : self.spnNom_Dsp_Idle23,
            'NOM_VEL_IDLE' : self.spnNom_Vel_Idle23,
            'NOM_ACCEL_IDLE' : self.spnNom_Accel_Idle23,
            'NOM_DSP_MEASURE' : self.spnNom_Dsp_Measure23,
            'NOM_VEL_MEASURE' : self.spnNom_Vel_Measure23,
            'NOM_ACCEL_MEASURE' : self.spnNom_Accel_Measure23,
            'NOM_LOAD' : self.spnNom_Load23,
            'NOM_POS_MEASURE' : self.spnNom_Pos_Measure23,
            'MICROSTEP' : self.spnMicrostep23},

            {'GEAR' : self.spnGear24,
            'PITCH' : self.spnPitch24,
            'NOM_TRAVEL' : self.spnNom_Travel24,
            'NOM_OMEGA' : self.spnNom_Omega24,
            'NOM_ACCEL_COEFF' : self.spnNom_Accel_Coeff24,
            'NOM_F1' : self.spnNom_F1_24,
            'NOM_F2' : self.spnNom_F2_24,
            'OVERLOAD_COEFF' : self.spnOverload_Coeff24,
            'DWELL' : self.spnDwell24,
            'N' : self.spnN24,
            'LOG_FREQ' : self.spnLog_Freq24,
            'MICROSTEP' : self.spnMicrostep24} )

    def save_ini(self, n_form):
        assert 0 < n_form < 5, "Номер формы (типа испытания) не входит в пределы 1..4 и равен " + str(n_form)
        # создать config из шаблона
        config = configparser.ConfigParser(strict=False) # strict=False - разрешение повторов имени ключа
        config.optionxform = str # имена ключей не будут приведены к нижнему регистру
        #config.read('BallScrewVCP_template.ini', encoding='utf-8')

        # добавить значения из интерфейса в объект config
        config['BALLSCREWPARAMS'] = {}
        for key, control in self.params_and_controls_dict[n_form-1].items():
            #DEBUG print "*** class name = ", control.__class__.__name__
            if (control.__class__.__name__ == 'QTimeEdit'):
                config['BALLSCREWPARAMS'][key] = str(QTime(0, 0, 0).secsTo(control.time()))
                config['BALLSCREWPARAMS'][key+'_MIN'] = str(QTime(0, 0, 0).secsTo(control.minimumTime()))
                config['BALLSCREWPARAMS'][key+'_MAX'] = str(QTime(0, 0, 0).secsTo(control.maximumTime()))
            else:
                config['BALLSCREWPARAMS'][key] = str(control.value())
                config['BALLSCREWPARAMS'][key+'_MIN'] = str(control.minimum())
                config['BALLSCREWPARAMS'][key+'_MAX'] = str(control.maximum())

        tmp_gear = self.params_and_controls_dict[n_form-1]['GEAR'].value();
        tmp_microstep = self.params_and_controls_dict[n_form-1]['MICROSTEP'].value()
        if n_form==2:
            config['BALLSCREWPARAMS']['SCALE'] = str(tmp_gear*tmp_microstep)
            config['BALLSCREWPARAMS']['SCALE_ENCODER'] = str(tmp_gear*tmp_microstep*20000.0/360.0)
        else:
            tmp_pitch = self.params_and_controls_dict[n_form-1]['PITCH'].value()
            config['BALLSCREWPARAMS']['SCALE'] = str(tmp_gear*tmp_microstep/tmp_pitch)
            config['BALLSCREWPARAMS']['SCALE_ENCODER'] = str(tmp_gear*tmp_microstep/tmp_pitch*20000.0/360.0)

        config['BALLSCREWPARAMS']['NAME'] = self.NAME
        config['BALLSCREWPARAMS']['TYPE'] = str(n_form) #TODO
        config['BALLSCREWPARAMS']['DATE'] = self.DATE
        config['BALLSCREWPARAMS']['MODEL'] = "TODO_need_for_Excel_file" #self.MODEL
        config['BALLSCREWPARAMS']['PART'] = self.PART
        print '***self.LOGFILE=', self.LOGFILE
        config['BALLSCREWPARAMS']['LOGFILE']= self.LOGFILE # "TempLogFile.log" # TODO вызывает ошибку строка в utf-8 с кириллицей

        rodos4_numbers_for_forms = [
            [0b01000,   # форма 3.1, перпендикулярный
             0b10000],  # форма 3.1, соосный
            [0b101000,  # форма 3.2, перпендикулярный
             0b110000], # форма 3.2, соосный
            [0b001010,  # форма 3.3, перпендикулярный
             0b010010], # форма 3.3, соосный
            [0b001010,  # форма 3.4, перпендикулярный
             0b010010]] # форма 3.4, соосный
        config['BALLSCREWPARAMS']['RODOS4_NUMBER']= (
            str(rodos4_numbers_for_forms[n_form-1][self.DRIVE]))

        if self.DRIVE==0:
            halfilename = 'BallScrewVCP_perp.hal'
        else:
            halfilename = 'BallScrewVCP_axial.hal'

        replacements = [
            #['{{MACHINE}}', 'BallScrewVCP' + str(n_form)],
            #['{{VCP}}', 'BallScrewVCP' + str(n_form)],
            #['{{DISPLAY}}', 'BallScrewVCP' + str(n_form)],
            #['{{POSTGUI_HALFILE}}', 'BallScrewVCP' + str(n_form) + '_postgui.hal'],
            ['{{BALL_SCREW_VCP_HAL}}', halfilename]]

        # открыть файл шаблона (свой у для каждой из форм 3.1-3.4)
        #f = open('BallScrewVCP_template.ini', 'rb')
        f = open('template_ini' + str(n_form) + '.txt', 'rb')
        filedata = f.read()
        f.close()
        for replacement in replacements:
            filedata = filedata.replace(replacement[0], replacement[1])
        # после разбивки на 4 отдельных файла
        # config['EMC'] = {}
        # config['EMC']['MACHINE'] = 'BallScrewVCP' + str(n_form)
        # config['HAL'] = {}
        # config['HAL']['POSTGUI_HALFILE'] = 'BallScrewVCP' + str(n_form)+'_postgui.hal'
        # config['DISPLAY'] = {}
        # config['DISPLAY']['VCP'] = 'BallScrewVCP' + str(n_form)
        # config['DISPLAY']['DISPLAY'] = 'qtvcp BallScrewVCP' + str(n_form)

        # запись заполненного config в ini-файл
        #.encode('utf-8')
        print '***self.NAME=', self.NAME

        iniFileName = "BallScrewVCP"+str(n_form)+".ini"
        # сначала записать фрагмент из шаблона с заменой {{ПАРАМЕТРОВ}}
        configfile = open(iniFileName, 'w')
        configfile.write(filedata)
        configfile.close()

        # добавление в конец файла секции [BALLSCREWVCP]
        # copyfile('BallScrewVCP_template.ini', self.NAME)
        configfile = codecs.open(iniFileName,'ab+','utf-8')
        config.write(configfile)
        configfile.close()

        #TODO обработка ошибок и исключений: нельзя открыть шаблон, нельзя открыть новый ini и пр.
        #TODO предупреждение о перезаписи
        return iniFileName

    def onBtnSaveIni(self):
        self.LOGFILE = self.edtFileSelect.text().encode('utf-8')
        print '***self.LOGFILE=', self.LOGFILE
        self.CNCINIFILENAME = self.save_ini(self.TYPE_N)
        return

    # тестовые кнопки для переключения страниц, удаляются/скрываются на релизе
    def onBtnTempShow1(self):
        self.stackedWidget.setCurrentIndex(0)
        pass

    def onBtnTempShow21(self):
        self.stackedWidget.setCurrentIndex(1)
        pass

    def onBtnTempShow22(self):
        self.stackedWidget.setCurrentIndex(2)
        pass

    def onBtnTempShow23(self):
        self.stackedWidget.setCurrentIndex(3)
        pass

    def onBtnTempShow24(self):
        self.stackedWidget.setCurrentIndex(4)
        pass

    # если используется PySide вместо PyQt
    # def load_ui(self):
    #     loader = QUiLoader()
    #     path = os.path.join(os.path.dirname(__file__), "BallScrew.ui")
    #     ui_file = QFile(path)
    #     ui_file.open(QFile.ReadOnly)
    #     loader.load(ui_file, self)
    #     ui_file.close()

if __name__ == "__main__":
    app = QApplication([])
    widget = BallScrew()
    widget.show()
    sys.exit(app.exec_())
