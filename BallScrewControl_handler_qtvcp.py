#!/usr/bin/python
# -*- coding: utf-8 -*-

############################
# **** IMPORT SECTION **** #
############################
import sys
import os
import linuxcnc
import openpyxl

from PyQt5 import QtCore, QtWidgets

from qtpyvcp.widgets.form_widgets.main_window import VCPMainWindow
#from qtvcp.widgets.mdi_line import MDILine as MDI_WIDGET
#from qtvcp.widgets.gcode_editor import GcodeEditor as GCODE
#from qtvcp.lib.keybindings import Keylookup
# from qtvcp.core import Status, Action

# Set up logging
#from qtvcp import logger
#LOG = logger.getLogger(__name__)

# Setup logging
from qtpyvcp.utilities import logger
LOG = logger.getLogger(__name__)

# Set the log level for this module
#LOG.setLevel(logger.INFO) # One of DEBUG, INFO, WARNING, ERROR, CRITICAL

###########################################
# **** INSTANTIATE LIBRARIES SECTION **** #
###########################################

#KEYBIND = Keylookup()
#STATUS = Status()
#ACTION = Action()

from qtpyvcp.plugins import getPlugin
STATUS = getPlugin('status')
STAT = STATUS.stat
STAT = linuxcnc.stat()

###################################
# **** HANDLER CLASS SECTION **** #
###################################

class BallScrewControlMainWindow(VCPMainWindow):

    ########################
    # **** INITIALIZE **** #
    ########################
    # widgets allows access to  widgets from the qtvcp files
    # at this point the widgets and hal pins are not instantiated
    # def __init__(self, halcomp,widgets,paths):
    def __init__(self, *args, **kwargs):
        super(BallScrewControlMainWindow, self).__init__(*args, **kwargs)
        self.hal = halcomp
        self.w = widgets
        self.PATHS = paths

    ##########################################
    # SPECIAL FUNCTIONS SECTION              #
    ##########################################

    # at this point:
    # the widgets are instantiated.
    # the HAL pins are built but HAL is not set ready
    # This is where you make HAL pins or initialize state of widgets etc
    def initialized__(self):
        try:
            wb = openpyxl.load_workbook(filename = database_file, read_only = True)
            print("*** OK openpyxl.load_workbook")
        except:
            # предупреждение в интерфейсе, что не удалось загрузить
            print("*** EXCEPTION openpyxl.load_workbook")
            return

        #TODO cmbModel.
        pass
        
    def onBtnNext(self):
        print "*** onBtnNext clicked"
        if(self.ui.cmbTestType.currentIndex()<4):
            self.ui.stackedWidget.setCurrentIndex(self.ui.cmbTestType.currentIndex()+1)
        else:
            self.ui.stackedWidget.setCurrentIndex(3)
        # self.close() только для случая многооконного интерфейса

################################
# required handler boiler code #
################################

def get_handlers(halcomp,widgets,paths):
     return [BallScrewControlMainWindow(halcomp,widgets,paths)]