import openpyxl
#from openpyxl import load_workbook
import os, sys
from PyQt5.QtWidgets import (QWidget, QHBoxLayout,
                             QLabel, QApplication,
                             QComboBox)

database_file = 'База данных испытаний.xlsx'

try:
    wb = openpyxl.load_workbook(filename = database_file, read_only = True)
except:
    print("Не удалось открыть файл базы данных \"", database_file, "\"")
    os._exit(1)

print(wb.sheetnames)

required_worksheets = ['Модели', 'Программы']
isDatabaseCorrect = True
for worksheet_name in required_worksheets:
    isDatabaseCorrect = isDatabaseCorrect and (worksheet_name in wb.sheetnames)
    if worksheet_name in wb.sheetnames:
        print("*Лист ", worksheet_name, " присутствует в БД")
    else:
        print("*Ошибка: ", worksheet_name, " отсутствует в БД")

if(not isDatabaseCorrect ):
    print("*** База данных некорректна")
    os._exit(1)

while

class Example(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):

        hbox = QHBoxLayout(self)
        combobox = QComboBox(self)
        combobox.addItems(wb.sheetnames)
        hbox.addWidget(combobox)
        self.setLayout(hbox)

        self.move(300, 200)
        self.setWindowTitle('Combobox from Excel')
        self.show()

if __name__ == '__main__':

    app = QApplication(sys.argv)
    ex = Example()
    sys.exit(app.exec_())